import os
import shutil
from datetime import datetime

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates

from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime
from sqlalchemy.orm import sessionmaker, declarative_base

from core.ocr import analyze_invoice
from core.pdf_generator import fill_voucher, merge_pdfs, create_advance_pdf

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


app = FastAPI()
templates = Jinja2Templates(directory="templates")

os.makedirs("data", exist_ok=True)
os.makedirs("uploads/invoices", exist_ok=True)
os.makedirs("uploads/vouchers", exist_ok=True)

engine = create_engine(
    "sqlite:///./data/budget.db",
    connect_args={"check_same_thread": False}
)

SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()


USERS = {
    "nazim": {"password": "nazim123", "role": "staff"},
    "maal": {"password": "maal123", "role": "staff"},
    "admin": {"password": "admin123", "role": "staff"},
}


class Budget(Base):
    __tablename__ = "budget"

    id = Column(Integer, primary_key=True)
    total = Column(Float, default=50000)


class Voucher(Base):
    __tablename__ = "vouchers"

    id = Column(Integer, primary_key=True)
    code = Column(String, unique=True)

    name = Column(String)
    department = Column(String)
    purpose = Column(String)
    amount = Column(Float)
    supplier = Column(String)

    voucher_type = Column(String, default="normal")
    advance_amount = Column(Float, default=0.0)
    receipt_status = Column(String, default="nicht_erforderlich")

    invoice_path = Column(String, nullable=True)
    voucher_pdf = Column(String, nullable=True)

    status = Column(String, default="entwurf")

    created_at = Column(DateTime, default=datetime.utcnow)
    approved_at = Column(DateTime, nullable=True)
    approved_by = Column(String, nullable=True)

    paid_at = Column(DateTime, nullable=True)
    paid_by = Column(String, nullable=True)


class History(Base):
    __tablename__ = "history"

    id = Column(Integer, primary_key=True)
    voucher_id = Column(Integer, nullable=True)
    action = Column(String)
    text = Column(String)
    person = Column(String, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)


def current_user(request: Request):
    username = request.cookies.get("username")
    role = request.cookies.get("role")

    if username in USERS and USERS[username]["role"] == role:
        return {"username": username, "role": role}

    return None


def require_staff(request: Request):
    return current_user(request)


def get_budget(db):
    budget = db.query(Budget).first()

    if not budget:
        budget = Budget(total=453867)
        db.add(budget)
        db.commit()
        db.refresh(budget)

    return budget


def add_history(db, action, text, voucher_id=None, person=None):
    log = History(
        voucher_id=voucher_id,
        action=action,
        text=text,
        person=person
    )
    db.add(log)
    db.commit()


def make_voucher_code(voucher_id: int):
    return f"V-2026-{voucher_id:04d}"


def parse_amount(value: str):
    try:
        return float(value.replace(".", "").replace(",", "."))
    except Exception:
        return 0.0


def budget_stats(db):
    budget = get_budget(db)

    reserved = sum(
        v.amount or 0
        for v in db.query(Voucher).filter(
            Voucher.status.in_(["genehmigt", "ausgezahlt", "rechnung_ausstehend", "abgeschlossen"])
        )
    )

    pending = sum(
        v.amount or 0
        for v in db.query(Voucher).filter(Voucher.status == "eingereicht")
    )

    paid = sum(
        v.amount or 0
        for v in db.query(Voucher).filter(
            Voucher.status.in_(["ausgezahlt", "rechnung_ausstehend", "abgeschlossen"])
        )
    )

    remaining = budget.total - reserved

    vouchers_count = db.query(Voucher).filter(Voucher.status != "entwurf").count()

    return budget, reserved, pending, paid, remaining, vouchers_count


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    db = SessionLocal()
    budget, reserved, pending, paid, remaining, vouchers_count = budget_stats(db)
    user = current_user(request)
    db.close()

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "budget": budget,
            "reserved": reserved,
            "pending": pending,
            "paid": paid,
            "remaining": remaining,
            "vouchers_count": vouchers_count,
            "user": user,
        }
    )


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"error": None})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user = USERS.get(username)

    if not user or user["password"] != password:
        return templates.TemplateResponse(
            request,
            "login.html",
            {"error": "Benutzername oder Passwort ist falsch."}
        )

    response = RedirectResponse("/", status_code=303)
    response.set_cookie("username", username)
    response.set_cookie("role", user["role"])
    return response


@app.get("/logout")
def logout():
    response = RedirectResponse("/", status_code=303)
    response.delete_cookie("username")
    response.delete_cookie("role")
    return response


@app.get("/create", response_class=HTMLResponse)
def create_page(request: Request):
    return templates.TemplateResponse(request, "create_voucher.html", {})

@app.get("/create-advance", response_class=HTMLResponse)
def create_advance_page(request: Request):
    return templates.TemplateResponse(
        request,
        "create_advance.html",
        {}
    )


@app.post("/create-advance")
def create_advance(
    request: Request,
    name: str = Form(...),
    member_id: str = Form(""),
    phone: str = Form(""),
    department: str = Form(""),
    advance_amount: str = Form(...),
    date: str = Form(""),
    purpose: str = Form(...)
):
    db = SessionLocal()

    amount = parse_amount(advance_amount)

    voucher = Voucher(
        code="TEMP",
        name=name,
        department=department,
        purpose=purpose,
        amount=amount,
        supplier="ADVANCE",
        voucher_type="advance",
        advance_amount=amount,
        receipt_status="ausstehend",
        status="eingereicht"
    )

    db.add(voucher)
    db.commit()
    db.refresh(voucher)

    voucher.code = make_voucher_code(voucher.id)

    db.commit()
    db.refresh(voucher)

    advance_pdf_path = f"uploads/vouchers/{voucher.code}_ADVANCE.pdf"

    advance_data = {
        "code": voucher.code,
        "date": date,
        "name": name,
        "member_id": member_id,
        "phone": phone,
        "department": department,
        "advance_amount": advance_amount,
        "purpose": purpose,
    }

    create_advance_pdf(
        advance_pdf_path,
        advance_data
    )

    voucher.voucher_pdf = advance_pdf_path

    db.commit()

    add_history(
        db,
        action="advance_beantragt",
        text=f"Advance {voucher.code} über {advance_amount} € wurde beantragt.",
        voucher_id=voucher.id,
        person=name
    )

    code = voucher.code

    db.close()

    return RedirectResponse(
        f"/submitted/{code}",
        status_code=303
    )
@app.post("/api/analyze-invoices")
async def analyze_invoices_api(invoices: list[UploadFile] = File(default=[])):
    results = []
    total_amount = 0.0
    supplier = ""
    iban = ""
    invoice_numbers = []
    date = ""

    for invoice in invoices:
        if not invoice or not invoice.filename:
            continue

        filename = invoice.filename.replace(" ", "_")
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        saved_path = f"uploads/invoices/{timestamp}_{filename}"

        with open(saved_path, "wb") as buffer:
            shutil.copyfileobj(invoice.file, buffer)

        try:
            analysis = analyze_invoice(saved_path)

            amount_text = analysis.get("amount", "")
            amount_value = parse_amount(amount_text)
            total_amount += amount_value

            if not supplier and analysis.get("supplier"):
                supplier = analysis.get("supplier")

            if not iban and analysis.get("iban"):
                iban = analysis.get("iban")

            if analysis.get("invoice_number"):
                invoice_numbers.append(analysis.get("invoice_number"))

            if not date and analysis.get("date"):
                date = analysis.get("date")

            results.append({
                "file": invoice.filename,
                "supplier": analysis.get("supplier", ""),
                "amount": analysis.get("amount", ""),
                "iban": analysis.get("iban", ""),
                "invoice_number": analysis.get("invoice_number", ""),
                "date": analysis.get("date", ""),
            })

        except Exception as e:
            results.append({
                "file": invoice.filename,
                "error": str(e)
            })

    total_text = (
        f"{total_amount:,.2f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )

    verwendungszweck = ""
    if invoice_numbers:
        verwendungszweck = "Rechnung " + ", ".join(invoice_numbers)

    return JSONResponse({
        "results": results,
        "total": total_text if total_amount > 0 else "",
        "supplier": supplier,
        "iban": iban,
        "date": date,
        "verwendungszweck": verwendungszweck
    })


@app.post("/create/finalize")
async def finalize_voucher(
    request: Request,

    invoices: list[UploadFile] = File(default=[]),

    scope: str = Form(""),
    datum: str = Form(""),

    name_amila: str = Form(""),
    id_amila: str = Form(""),

    name_antrag: str = Form(...),
    id_antrag: str = Form(""),

    tel_antrag: str = Form(""),
    abteilung: str = Form(""),

    zweck: str = Form(""),
    teilnehmer: str = Form(""),

    advance_erhalten: str = Form(""),
    advance_betrag: str = Form(""),

    pos1_betrag: str = Form(""),
    pos2_betrag: str = Form(""),
    pos3_betrag: str = Form(""),
    pos4_betrag: str = Form(""),
    pos5_betrag: str = Form(""),
    pos6_betrag: str = Form(""),
    pos7_betrag: str = Form(""),
    pos8_betrag: str = Form(""),
    pos9_betrag: str = Form(""),
    pos10_betrag: str = Form(""),

    total: str = Form(""),

    supplier: str = Form(""),
    verwendungszweck: str = Form(""),
    iban: str = Form(""),

    signature_applicant: str = Form(""),
    signature_amila: str = Form("")
):
    db = SessionLocal()

    amount = parse_amount(total)
    adv_amount = parse_amount(advance_betrag)

    is_advance = advance_erhalten == "Ja" or adv_amount > 0

    voucher = Voucher(
        code="TEMP",
        name=name_antrag,
        department=abteilung,
        purpose=zweck,
        amount=amount,
        supplier=supplier,
        voucher_type="advance" if is_advance else "normal",
        advance_amount=adv_amount,
        receipt_status="ausstehend" if is_advance and not invoices else "vorhanden" if is_advance else "nicht_erforderlich",
        status="entwurf"
    )

    db.add(voucher)
    db.commit()
    db.refresh(voucher)

    voucher.code = make_voucher_code(voucher.id)
    db.commit()
    db.refresh(voucher)

    saved_invoice_paths = []

    for invoice in invoices:
        if invoice and invoice.filename:
            filename = invoice.filename.replace(" ", "_")
            save_path = f"uploads/invoices/{voucher.code}_{filename}"

            with open(save_path, "wb") as buffer:
                shutil.copyfileobj(invoice.file, buffer)

            saved_invoice_paths.append(save_path)

    if saved_invoice_paths:
        voucher.invoice_path = ";".join(saved_invoice_paths)

    voucher_pdf_path = f"uploads/vouchers/{voucher.code}.pdf"

    data = {
        "scope": scope,
        "datum": datum,

        "name_amila": name_amila,
        "id_amila": id_amila,

        "name_antrag": name_antrag,
        "id_antrag": id_antrag,

        "tel_antrag": tel_antrag,
        "abteilung": abteilung,
        "zweck": zweck,
        "teilnehmer": teilnehmer,

        "advance_erhalten": advance_erhalten,
        "advance_betrag": advance_betrag,

        "pos1_betrag": pos1_betrag,
        "pos2_betrag": pos2_betrag,
        "pos3_betrag": pos3_betrag,
        "pos4_betrag": pos4_betrag,
        "pos5_betrag": pos5_betrag,
        "pos6_betrag": pos6_betrag,
        "pos7_betrag": pos7_betrag,
        "pos8_betrag": pos8_betrag,
        "pos9_betrag": pos9_betrag,
        "pos10_betrag": pos10_betrag,

        "total": total,
        "supplier": supplier,
        "verwendungszweck": verwendungszweck,
        "iban": iban,

        "signature_applicant": signature_applicant,
        "signature_amila": signature_amila,
    }

    fill_voucher(voucher_pdf_path, data)

    final_pdf = f"uploads/vouchers/{voucher.code}_FULL.pdf"

    if saved_invoice_paths:
        current_pdf = voucher_pdf_path

        for index, invoice_path in enumerate(saved_invoice_paths):
            temp_output = f"uploads/vouchers/{voucher.code}_TEMP_{index}.pdf"
            merge_pdfs(current_pdf, invoice_path, temp_output)
            current_pdf = temp_output

        os.replace(current_pdf, final_pdf)
        voucher.voucher_pdf = final_pdf
    else:
        voucher.voucher_pdf = voucher_pdf_path

    db.commit()

    add_history(
        db,
        action="entwurf",
        text=f"Voucher {voucher.code} wurde als Vorschau erstellt.",
        voucher_id=voucher.id,
        person=name_antrag
    )

    code = voucher.code
    db.close()

    return RedirectResponse(f"/preview/{code}", status_code=303)


@app.get("/preview/{code}", response_class=HTMLResponse)
def preview_page(request: Request, code: str):
    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.code == code).first()
    db.close()

    if not voucher:
        return HTMLResponse("<h1>Voucher nicht gefunden</h1>", status_code=404)

    return HTMLResponse(f"""
    <!DOCTYPE html>
    <html lang="de">
    <head>
        <meta charset="UTF-8">
        <title>Voucher Vorschau</title>
        <style>
            body {{ font-family: Arial, sans-serif; background:#e8edf2; margin:0; padding:30px; }}
            .header {{ background:#1b2d42; color:white; padding:24px; border-radius:12px; margin-bottom:20px; }}
            .card {{ background:white; padding:24px; border-radius:12px; border:1px solid #c8d0da; max-width:900px; margin-bottom:20px; }}
            a {{ display:inline-block; padding:12px 18px; border-radius:8px; text-decoration:none; font-weight:bold; margin-right:10px; margin-top:10px; }}
            .primary {{ background:#1a5276; color:white; }}
            .secondary {{ background:#7f8c8d; color:white; }}
        </style>
    </head>
    <body>

    <div class="header">
        <h1>Voucher Vorschau</h1>
        <p>Bitte öffne die PDF und prüfe alles. Erst danach endgültig einreichen.</p>
    </div>

    <div class="card">
        <h2>{voucher.code}</h2>
        <p><b>Antragsteller:</b> {voucher.name}</p>
        <p><b>Betrag:</b> {voucher.amount:.2f} €</p>
        <p><b>Typ:</b> {voucher.voucher_type}</p>
        <p><b>Belegstatus:</b> {voucher.receipt_status}</p>
        <p><b>Status:</b> Vorschau / Entwurf</p>

        <a class="primary" href="/pdf/{voucher.code}" target="_blank">PDF öffnen und prüfen</a>
        <a class="secondary" href="/create">Zurück / neu erstellen</a>
        <a class="primary" href="/submit/{voucher.code}">Jetzt endgültig einreichen</a>
    </div>

    </body>
    </html>
    """)


@app.get("/pdf/{code}")
def open_pdf(code: str):
    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.code == code).first()
    db.close()

    if not voucher or not voucher.voucher_pdf or not os.path.exists(voucher.voucher_pdf):
        return HTMLResponse("<h1>PDF nicht gefunden</h1>", status_code=404)

    return FileResponse(
        voucher.voucher_pdf,
        media_type="application/pdf",
        filename=os.path.basename(voucher.voucher_pdf)
    )


@app.get("/submit/{code}")
def submit_voucher(code: str):
    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.code == code).first()

    if not voucher:
        db.close()
        return HTMLResponse("<h1>Voucher nicht gefunden</h1>", status_code=404)

    voucher.status = "eingereicht"
    db.commit()

    add_history(
        db,
        action="eingereicht",
        text=f"Voucher {voucher.code} wurde endgültig eingereicht.",
        voucher_id=voucher.id,
        person=voucher.name
    )

    db.close()

    return RedirectResponse(f"/submitted/{code}", status_code=303)


@app.get("/submitted/{code}", response_class=HTMLResponse)
def submitted_page(request: Request, code: str):
    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.code == code).first()
    db.close()

    return templates.TemplateResponse(
        request,
        "submitted.html",
        {"voucher": voucher}
    )


@app.get("/track", response_class=HTMLResponse)
def track_page(request: Request):
    return templates.TemplateResponse(
        request,
        "track.html",
        {"voucher": None, "error": None}
    )


@app.post("/track", response_class=HTMLResponse)
def track_voucher(
    request: Request,
    code: str = Form(...),
    name: str = Form(...)
):
    db = SessionLocal()

    voucher = (
        db.query(Voucher)
        .filter(Voucher.code == code.strip())
        .filter(Voucher.name == name.strip())
        .filter(Voucher.status != "entwurf")
        .first()
    )

    db.close()

    if not voucher:
        return templates.TemplateResponse(
            request,
            "track.html",
            {
                "voucher": None,
                "error": "Kein eingereichter Voucher mit diesem Code und Namen gefunden."
            }
        )

    return templates.TemplateResponse(
        request,
        "track.html",
        {"voucher": voucher, "error": None}
    )


@app.get("/review", response_class=HTMLResponse)
def review_page(request: Request):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    vouchers = (
        db.query(Voucher)
        .filter(Voucher.status == "eingereicht")
        .order_by(Voucher.created_at.desc())
        .all()
    )

    db.close()

    return templates.TemplateResponse(
        request,
        "review.html",
        {"vouchers": vouchers, "user": user}
    )


@app.get("/approve/{voucher_id}")
def approve_voucher(request: Request, voucher_id: int):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()

    if voucher and voucher.status == "eingereicht":
        voucher.status = "genehmigt"
        voucher.approved_at = datetime.utcnow()
        voucher.approved_by = user["username"]
        db.commit()

        add_history(
            db,
            action="genehmigt",
            text=f"Voucher {voucher.code} wurde genehmigt.",
            voucher_id=voucher.id,
            person=user["username"]
        )

    db.close()
    return RedirectResponse("/review", status_code=303)


@app.get("/reject/{voucher_id}")
def reject_voucher(request: Request, voucher_id: int):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()

    if voucher and voucher.status == "eingereicht":
        voucher.status = "abgelehnt"
        db.commit()

        add_history(
            db,
            action="abgelehnt",
            text=f"Voucher {voucher.code} wurde abgelehnt.",
            voucher_id=voucher.id,
            person=user["username"]
        )

    db.close()
    return RedirectResponse("/review", status_code=303)


@app.get("/pay/{voucher_id}")
def pay_voucher(request: Request, voucher_id: int):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()

    if voucher and voucher.status == "genehmigt":
        voucher.paid_at = datetime.utcnow()
        voucher.paid_by = user["username"]

        if voucher.voucher_type == "advance" and voucher.receipt_status == "ausstehend":
            voucher.status = "rechnung_ausstehend"
        else:
            voucher.status = "ausgezahlt"

        db.commit()

        add_history(
            db,
            action="ausgezahlt",
            text=f"Voucher {voucher.code} wurde als ausgezahlt markiert.",
            voucher_id=voucher.id,
            person=user["username"]
        )

    db.close()
    return RedirectResponse("/vouchers", status_code=303)


@app.get("/upload-receipt/{voucher_id}", response_class=HTMLResponse)
def upload_receipt_page(request: Request, voucher_id: int):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    return HTMLResponse(f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Rechnung nachreichen</title>
    </head>
    <body style="font-family:Arial; background:#e8edf2; padding:30px;">
        <div style="background:white; padding:25px; border-radius:12px; max-width:600px;">
            <h1>Rechnung nachreichen</h1>
            <form method="post" action="/upload-receipt/{voucher_id}" enctype="multipart/form-data">
                <label>Rechnung(en) hochladen</label><br><br>
                <input type="file" name="receipts" multiple accept="application/pdf"><br><br>
                <button type="submit">Hochladen und abschließen</button>
            </form>
        </div>
    </body>
    </html>
    """)


@app.post("/upload-receipt/{voucher_id}")
async def upload_receipt(
    request: Request,
    voucher_id: int,
    receipts: list[UploadFile] = File(default=[])
):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()

    if not voucher:
        db.close()
        return HTMLResponse("<h1>Voucher nicht gefunden</h1>", status_code=404)

    saved_paths = []
    final_receipt_total = 0.0

    for receipt in receipts:
        if receipt and receipt.filename:
            filename = receipt.filename.replace(" ", "_")
            save_path = f"uploads/invoices/{voucher.code}_NACHGEREICHT_{filename}"

            with open(save_path, "wb") as buffer:
                shutil.copyfileobj(receipt.file, buffer)

            saved_paths.append(save_path)

            try:
                analysis = analyze_invoice(save_path)
                final_receipt_total += parse_amount(analysis.get("amount", ""))
            except Exception:
                pass

    if not saved_paths:
        db.close()
        return HTMLResponse("<h1>Keine Rechnung hochgeladen</h1>", status_code=400)

    old_paths = voucher.invoice_path.split(";") if voucher.invoice_path else []
    voucher.invoice_path = ";".join(old_paths + saved_paths)

    if not voucher.voucher_pdf or not os.path.exists(voucher.voucher_pdf):
        advance_pdf_path = f"uploads/vouchers/{voucher.code}_ADVANCE.pdf"

        create_advance_pdf(
            advance_pdf_path,
            {
                "code": voucher.code,
                "date": voucher.created_at.strftime("%d.%m.%Y"),
                "name": voucher.name,
                "member_id": "",
                "phone": "",
                "department": voucher.department,
                "advance_amount": f"{voucher.advance_amount:.2f}",
                "purpose": voucher.purpose,
            }
        )

        voucher.voucher_pdf = advance_pdf_path
        db.commit()

    current_pdf = voucher.voucher_pdf

    for index, receipt_path in enumerate(saved_paths):
        temp_output = f"uploads/vouchers/{voucher.code}_MERGE_{index}.pdf"

        merge_pdfs(
            current_pdf,
            receipt_path,
            temp_output
        )

        current_pdf = temp_output

    final_pdf = f"uploads/vouchers/{voucher.code}_FINAL.pdf"

    if os.path.exists(final_pdf):
        os.remove(final_pdf)

    os.replace(current_pdf, final_pdf)

    voucher.voucher_pdf = final_pdf

    difference = voucher.advance_amount - final_receipt_total

    if difference > 0:
        voucher.receipt_status = f"rueckgabe_offen:{difference:.2f}"
        voucher.status = "abrechnung_offen"
        history_text = f"Rechnungen über {final_receipt_total:.2f} € nachgereicht. Rückgabe offen: {difference:.2f} €."

    elif difference < 0:
        extra = abs(difference)
        voucher.receipt_status = f"nachzahlung_offen:{extra:.2f}"
        voucher.status = "abrechnung_offen"
        history_text = f"Rechnungen über {final_receipt_total:.2f} € nachgereicht. Nachzahlung offen: {extra:.2f} €."

    else:
        voucher.receipt_status = "abgerechnet"
        voucher.status = "abgeschlossen"
        history_text = f"Advance vollständig abgerechnet. Rechnungen: {final_receipt_total:.2f} €."

    db.commit()

    add_history(
        db,
        action="advance_abrechnung",
        text=history_text,
        voucher_id=voucher.id,
        person=user["username"]
    )

    db.close()

    return RedirectResponse("/advances", status_code=303)

@app.get("/settle-advance", response_class=HTMLResponse)
def settle_advance_page(request: Request):

    return templates.TemplateResponse(
        request,
        "settle_advance.html",
        {
            "error": None
        }
    )


@app.post("/settle-advance", response_class=HTMLResponse)
async def settle_advance(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    receipts: list[UploadFile] = File(default=[])
):
    db = SessionLocal()

    voucher = (
        db.query(Voucher)
        .filter(Voucher.code == code.strip())
        .filter(Voucher.name == name.strip())
        .filter(Voucher.voucher_type == "advance")
        .first()
    )

    if not voucher:

        db.close()

        return templates.TemplateResponse(
            request,
            "settle_advance.html",
            {
                "error": "Advance nicht gefunden."
            }
        )

    if voucher.status not in [
        "rechnung_ausstehend",
        "abrechnung_offen",
        "ausgezahlt"
    ]:

        db.close()

        return templates.TemplateResponse(
            request,
            "settle_advance.html",
            {
                "error": "Dieser Advance kann aktuell nicht abgerechnet werden."
            }
        )

    saved_paths = []
    final_receipt_total = 0.0

    for receipt in receipts:

        if receipt and receipt.filename:

            filename = receipt.filename.replace(" ", "_")

            save_path = (
                f"uploads/invoices/"
                f"{voucher.code}_PUBLIC_{filename}"
            )

            with open(save_path, "wb") as buffer:
                shutil.copyfileobj(receipt.file, buffer)

            saved_paths.append(save_path)

            try:

                analysis = analyze_invoice(save_path)

                amount = parse_amount(
                    analysis.get("amount", "")
                )

                final_receipt_total += amount

            except Exception:
                pass

    if not saved_paths:

        db.close()

        return templates.TemplateResponse(
            request,
            "settle_advance.html",
            {
                "error": "Keine Rechnungen hochgeladen."
            }
        )

    old_paths = (
        voucher.invoice_path.split(";")
        if voucher.invoice_path
        else []
    )

    voucher.invoice_path = ";".join(
        old_paths + saved_paths
    )

    current_pdf = voucher.voucher_pdf

    for index, receipt_path in enumerate(saved_paths):

        temp_output = (
            f"uploads/vouchers/"
            f"{voucher.code}_PUBLIC_{index}.pdf"
        )

        merge_pdfs(
            current_pdf,
            receipt_path,
            temp_output
        )

        current_pdf = temp_output

    final_pdf = (
        f"uploads/vouchers/"
        f"{voucher.code}_FINAL.pdf"
    )

    if os.path.exists(final_pdf):
        os.remove(final_pdf)

    os.replace(
        current_pdf,
        final_pdf
    )

    voucher.voucher_pdf = final_pdf

    difference = (
        voucher.advance_amount
        - final_receipt_total
    )

    if difference > 0:

        voucher.receipt_status = (
            f"rueckgabe_offen:{difference:.2f}"
        )

        voucher.status = "abrechnung_offen"

        history_text = (
            f"Öffentliche Advance-Abrechnung. "
            f"Rechnungen: {final_receipt_total:.2f} €. "
            f"Rückgabe offen: {difference:.2f} €."
        )

    elif difference < 0:

        extra = abs(difference)

        voucher.receipt_status = (
            f"nachzahlung_offen:{extra:.2f}"
        )

        voucher.status = "abrechnung_offen"

        history_text = (
            f"Öffentliche Advance-Abrechnung. "
            f"Rechnungen: {final_receipt_total:.2f} €. "
            f"Nachzahlung offen: {extra:.2f} €."
        )

    else:

        voucher.receipt_status = "abgerechnet"
        voucher.status = "abgeschlossen"

        history_text = (
            f"Advance vollständig öffentlich abgerechnet."
        )

    db.commit()

    add_history(
        db,
        action="public_advance_abrechnung",
        text=history_text,
        voucher_id=voucher.id,
        person=name
    )

    db.close()

    return RedirectResponse(
        f"/track?success={voucher.code}",
        status_code=303
    )

@app.get("/advances", response_class=HTMLResponse)
def advances_page(request: Request):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    vouchers = (
        db.query(Voucher)
        .filter(Voucher.voucher_type == "advance")
        .filter(Voucher.status != "entwurf")
        .order_by(Voucher.created_at.desc())
        .all()
    )

    db.close()

    rows = ""

    for v in vouchers:
        upload_button = ""
        if v.receipt_status == "ausstehend":
            upload_button = f'<a href="/upload-receipt/{v.id}">Rechnung nachreichen</a>'

        rows += f"""
        <div style="background:white; padding:18px; border-radius:10px; margin-bottom:12px;">
            <b>{v.code}</b><br>
            Antragsteller: {v.name}<br>
            Zweck: {v.purpose}<br>
            Betrag: {v.amount:.2f} €<br>
            Advance: {v.advance_amount:.2f} €<br>
            Status: <b>{v.status}</b><br>
            Belegstatus: <b>{v.receipt_status}</b><br>
            <a href="/pdf/{v.code}" target="_blank">PDF öffnen</a>
            {upload_button}
        </div>
        """

    return HTMLResponse(f"""
    <html>
    <head><meta charset="UTF-8"><title>Advances</title></head>
    <body style="font-family:Arial; background:#e8edf2; padding:30px;">
        <div style="background:#1b2d42; color:white; padding:20px; border-radius:12px; margin-bottom:20px;">
            <h1>Advance Übersicht</h1>
            <a href="/" style="color:#7fb3d3;">Dashboard</a>
        </div>
        {rows if rows else "<p>Keine Advances vorhanden.</p>"}
    </body>
    </html>
    """)


@app.get("/history", response_class=HTMLResponse)
def history_page(request: Request):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()
    logs = db.query(History).order_by(History.created_at.desc()).all()
    db.close()

    return templates.TemplateResponse(
        request,
        "history.html",
        {"logs": logs, "user": user}
    )


@app.get("/vouchers", response_class=HTMLResponse)
def vouchers_page(request: Request):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    vouchers = (
        db.query(Voucher)
        .filter(Voucher.status != "entwurf")
        .order_by(Voucher.created_at.desc())
        .all()
    )

    db.close()

    return templates.TemplateResponse(
        request,
        "vouchers.html",
        {"vouchers": vouchers, "user": user}
    )
@app.get("/reset-test-data")
def reset_test_data(request: Request):
    db = SessionLocal()

    db.query(History).delete()
    db.query(Voucher).delete()
    db.query(Budget).delete()

    budget = Budget(total=50000)
    db.add(budget)

    db.commit()
    db.close()

    return RedirectResponse("/", status_code=303)
@app.get("/files", response_class=HTMLResponse)
def files_page(
    request: Request,
    q: str = "",
    typ: str = "",
    status: str = ""
):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    query = db.query(Voucher).filter(Voucher.status != "entwurf")

    if q:
        search = f"%{q}%"
        query = query.filter(
            (Voucher.code.like(search)) |
            (Voucher.name.like(search)) |
            (Voucher.purpose.like(search))
        )

    if typ:
        query = query.filter(Voucher.voucher_type == typ)

    if status:
        query = query.filter(Voucher.status == status)

    vouchers = query.order_by(Voucher.created_at.desc()).all()

    db.close()

    rows = ""

    for v in vouchers:
        pdf_link = (
            f'<a class="btn" href="/pdf/{v.code}" target="_blank">PDF öffnen</a>'
            if v.voucher_pdf
            else "<span class='muted'>Keine PDF</span>"
        )

        rows += f"""
        <tr>
            <td><b>{v.code}</b></td>
            <td>{v.name}</td>
            <td>{v.purpose or "-"}</td>
            <td><span class="badge type">{v.voucher_type}</span></td>
            <td><span class="badge status">{v.status}</span></td>
            <td>{v.receipt_status or "-"}</td>
            <td>{v.amount:.2f} €</td>
            <td>{pdf_link}</td>
        </tr>
        """

    return HTMLResponse(f"""
    <!DOCTYPE html>
    <html lang="de">
    <head>
        <meta charset="UTF-8">
        <title>Dateimanager</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background:#f3f6fb;
                margin:0;
                padding:30px;
                color:#132238;
            }}

            .header {{
                background:#1b2d42;
                color:white;
                padding:24px;
                border-radius:14px;
                margin-bottom:20px;
            }}

            .header a {{
                color:#7fb3d3;
                font-weight:bold;
                text-decoration:none;
                margin-right:16px;
            }}

            .card {{
                background:white;
                padding:24px;
                border-radius:14px;
                border:1px solid #d8e0ea;
                margin-bottom:20px;
            }}

            .filters {{
                display:grid;
                grid-template-columns:2fr 1fr 1fr auto;
                gap:12px;
                align-items:end;
            }}

            input, select {{
                width:100%;
                padding:11px;
                border:1px solid #b0bbc8;
                border-radius:8px;
            }}

            label {{
                font-weight:bold;
                display:block;
                margin-bottom:6px;
            }}

            button {{
                padding:12px 18px;
                background:#1a5276;
                color:white;
                border:0;
                border-radius:8px;
                font-weight:bold;
                cursor:pointer;
            }}

            table {{
                width:100%;
                border-collapse:collapse;
                margin-top:10px;
            }}

            th, td {{
                padding:13px;
                border-bottom:1px solid #e1e7ef;
                text-align:left;
                font-size:14px;
            }}

            th {{
                background:#f3f6fb;
                color:#1b2d42;
            }}

            .btn {{
                background:#1a5276;
                color:white;
                padding:8px 12px;
                border-radius:8px;
                text-decoration:none;
                font-weight:bold;
            }}

            .badge {{
                padding:5px 10px;
                border-radius:999px;
                font-weight:bold;
                font-size:12px;
                display:inline-block;
            }}

            .type {{
                background:#e0ecff;
                color:#1d4ed8;
            }}

            .status {{
                background:#ecfdf5;
                color:#047857;
            }}

            .muted {{
                color:#777;
            }}

            @media(max-width:900px) {{
                .filters {{
                    grid-template-columns:1fr;
                }}

                table {{
                    font-size:12px;
                }}
            }}
        </style>
    </head>

    <body>

        <div class="header">
            <h1>Dateimanager</h1>
            <p>Alle Voucher, Advances und PDFs zentral durchsuchen.</p>
            <a href="/">Dashboard</a>
            <a href="/vouchers">Alle Voucher</a>
            <a href="/advances">Advances</a>
            <a href="/history">Historie</a>
        </div>

        <div class="card">
            <form method="get" action="/files" class="filters">
                <div>
                    <label>Suchen</label>
                    <input type="text" name="q" value="{q}" placeholder="Code, Name oder Zweck">
                </div>

                <div>
                    <label>Typ</label>
                    <select name="typ">
                        <option value="">Alle</option>
                        <option value="normal" {"selected" if typ == "normal" else ""}>Normal</option>
                        <option value="advance" {"selected" if typ == "advance" else ""}>Advance</option>
                    </select>
                </div>

                <div>
                    <label>Status</label>
                    <select name="status">
                        <option value="">Alle</option>
                        <option value="eingereicht" {"selected" if status == "eingereicht" else ""}>Eingereicht</option>
                        <option value="genehmigt" {"selected" if status == "genehmigt" else ""}>Genehmigt</option>
                        <option value="ausgezahlt" {"selected" if status == "ausgezahlt" else ""}>Ausgezahlt</option>
                        <option value="rechnung_ausstehend" {"selected" if status == "rechnung_ausstehend" else ""}>Rechnung ausstehend</option>
                        <option value="abrechnung_offen" {"selected" if status == "abrechnung_offen" else ""}>Abrechnung offen</option>
                        <option value="abgeschlossen" {"selected" if status == "abgeschlossen" else ""}>Abgeschlossen</option>
                        <option value="abgelehnt" {"selected" if status == "abgelehnt" else ""}>Abgelehnt</option>
                    </select>
                </div>

                <button type="submit">Filtern</button>
            </form>
        </div>

        <div class="card">
            <table>
                <thead>
                    <tr>
                        <th>Code</th>
                        <th>Name</th>
                        <th>Zweck</th>
                        <th>Typ</th>
                        <th>Status</th>
                        <th>Belegstatus</th>
                        <th>Betrag</th>
                        <th>Datei</th>
                    </tr>
                </thead>
                <tbody>
                    {rows if rows else "<tr><td colspan='8'>Keine Dateien gefunden.</td></tr>"}
                </tbody>
            </table>
        </div>

    </body>
    </html>
    """)
@app.get("/export-excel")
def export_excel(request: Request):
    user = require_staff(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    db = SessionLocal()

    vouchers = (
        db.query(Voucher)
        .filter(Voucher.status != "entwurf")
        .order_by(Voucher.created_at.desc())
        .all()
    )

    logs = (
        db.query(History)
        .order_by(History.created_at.desc())
        .all()
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Alle Voucher"

    headers = [
        "Code",
        "Name",
        "Typ",
        "Status",
        "Belegstatus",
        "Zweck",
        "Abteilung",
        "Betrag",
        "Advance Betrag",
        "Lieferant",
        "Erstellt am",
        "Genehmigt von",
        "Ausgezahlt von",
        "PDF Pfad",
    ]

    ws.append(headers)

    for v in vouchers:
        ws.append([
            v.code,
            v.name,
            v.voucher_type,
            v.status,
            v.receipt_status,
            v.purpose,
            v.department,
            v.amount,
            v.advance_amount,
            v.supplier,
            v.created_at.strftime("%d.%m.%Y %H:%M") if v.created_at else "",
            v.approved_by or "",
            v.paid_by or "",
            v.voucher_pdf or "",
        ])

    adv = wb.create_sheet("Advances")

    adv_headers = [
        "Code",
        "Name",
        "Zweck",
        "Advance Betrag",
        "Status",
        "Belegstatus",
        "Rückgabe/Nachzahlung",
        "PDF Pfad",
    ]

    adv.append(adv_headers)

    for v in vouchers:
        if v.voucher_type == "advance":
            balance = ""

            if v.receipt_status and ":" in v.receipt_status:
                balance = v.receipt_status.replace(":", " ")

            adv.append([
                v.code,
                v.name,
                v.purpose,
                v.advance_amount,
                v.status,
                v.receipt_status,
                balance,
                v.voucher_pdf or "",
            ])

    hist = wb.create_sheet("Historie")

    hist_headers = [
        "Datum",
        "Voucher ID",
        "Aktion",
        "Beschreibung",
        "Person",
    ]

    hist.append(hist_headers)

    for h in logs:
        hist.append([
            h.created_at.strftime("%d.%m.%Y %H:%M") if h.created_at else "",
            h.voucher_id,
            h.action,
            h.text,
            h.person or "",
        ])

    summary = wb.create_sheet("Übersicht", 0)

    total_amount = sum(v.amount or 0 for v in vouchers)
    total_advance = sum(v.advance_amount or 0 for v in vouchers if v.voucher_type == "advance")
    open_advances = len([v for v in vouchers if v.voucher_type == "advance" and v.status in ["rechnung_ausstehend", "abrechnung_offen"]])

    summary.append(["Ijtema Budget System Export"])
    summary.append([])
    summary.append(["Anzahl Voucher", len(vouchers)])
    summary.append(["Gesamtsumme Voucher", total_amount])
    summary.append(["Advance Summe", total_advance])
    summary.append(["Offene Advances", open_advances])
    summary.append(["Export erstellt am", datetime.utcnow().strftime("%d.%m.%Y %H:%M")])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1B2D42")
            cell.alignment = Alignment(horizontal="center")

        for col in sheet.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                except Exception:
                    pass

            width = min(max_length + 3, 45)
            sheet.column_dimensions[get_column_letter(column)].width = width

    os.makedirs("exports", exist_ok=True)

    export_path = "exports/ijtema_budget_export.xlsx"

    wb.save(export_path)

    db.close()

    return FileResponse(
        export_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="ijtema_budget_export.xlsx"
    )